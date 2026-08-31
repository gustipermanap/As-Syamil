from datetime import date, timedelta, time
from decimal import Decimal

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from akademik.models import Absensi, JadwalSlot, KeanggotaanRB, KitabAtauMapel, MapelRB, PengampuRB, Pertemuan, RombonganBelajar, RuangBelajar
from akademik.services import salin_rb, streak_alpa
from kesiswaan.models import Izin, Pegawai, Santri
from lembaga.factories import buat_santri, buat_user, data_dasar
from lembaga.models import Pengaturan, TahunAjaran
from ppdb.models import GelombangPPDB
from WebApp.models import Pendaftaran
from WebApp.tests import _ppdb_payload


class DaftarFilterTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.data = data_dasar()
        cls.tu = buat_user('tu_daftar', 'tata_usaha')
        cls.s1 = buat_santri('NISFLT1', 'Ahmad Filter', '3601010101010101')
        cls.s2 = buat_santri('NISFLT2', 'Budi Lulus', '3601010101010102', status='lulus')
        for i in range(30):
            buat_santri(f'NISPAG{i:03d}', f'Paginas {i:03d}', f'3601990101010{i:03d}')

    def setUp(self):
        self.client.login(username='tu_daftar', password='sandi123')

    def test_filter_nama_dan_status(self):
        r = self.client.get(reverse('kesiswaan:santri'), {'q': 'Ahmad Filter'})
        self.assertContains(r, 'Ahmad Filter')
        self.assertNotContains(r, 'Budi Lulus')
        r2 = self.client.get(reverse('kesiswaan:santri'), {'status': 'lulus'})
        self.assertContains(r2, 'Budi Lulus')
        self.assertNotContains(r2, 'Ahmad Filter')

    def test_pagination(self):
        r = self.client.get(reverse('kesiswaan:santri'))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Halaman 1')
        r2 = self.client.get(reverse('kesiswaan:santri'), {'page': 2})
        self.assertEqual(r2.status_code, 200)
        self.assertContains(r2, 'Halaman 2')

    def test_ekspor_excel_mengikuti_filter(self):
        from io import BytesIO
        from openpyxl import load_workbook
        r = self.client.get(reverse('kesiswaan:santri'), {'q': 'Ahmad Filter', 'ekspor': 'xlsx'})
        self.assertEqual(r.status_code, 200)
        self.assertEqual(
            r['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        wb = load_workbook(BytesIO(r.content))
        teks = ' '.join(
            str(v or '') for row in wb.active.iter_rows(values_only=True) for v in row
        )
        self.assertIn('Ahmad Filter', teks)
        self.assertNotIn('Budi Lulus', teks)

    def test_checkbox_aksi_massal(self):
        r = self.client.post(
            reverse('kesiswaan:santri'),
            {'pilih': [str(self.s1.pk)], 'aksi_massal': 'lulus'},
        )
        self.assertEqual(r.status_code, 302)
        self.s1.refresh_from_db()
        self.assertEqual(self.s1.status, 'lulus')

    def test_toolbar_tampil(self):
        r = self.client.get(reverse('kesiswaan:santri'))
        self.assertContains(r, 'Ekspor Excel')
        self.assertContains(r, 'Filter lanjutan')
        self.assertContains(r, 'pilih-semua')
        self.assertContains(r, 'Aksi untuk baris terpilih')


class PPDBKuotaTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        data_dasar()
        cls.tu = buat_user('tu_kuota', 'tata_usaha')

    def test_kuota_penuh_menjadi_cadangan_dan_tutup(self):
        now = timezone.now()
        g = GelombangPPDB.objects.create(
            nama='Kuota 1',
            mulai=now - timedelta(days=1),
            selesai=now + timedelta(days=10),
            kuota=1,
            status=GelombangPPDB.DIBUKA,
        )
        self.client.post(reverse('pendaftaran'), _ppdb_payload(nama_lengkap='Calon Satu'))
        self.client.post(reverse('pendaftaran'), _ppdb_payload(nama_lengkap='Calon Dua', nik='3601010101010999', nisn='1234567891'))
        p1 = Pendaftaran.objects.get(nama_lengkap='Calon Satu')
        p2 = Pendaftaran.objects.get(nama_lengkap='Calon Dua')
        self.client.login(username='tu_kuota', password='sandi123')
        self.client.get(reverse('ppdb:ubah_status', args=[p1.pk, 'diterima']))
        p1.refresh_from_db()
        g.refresh_from_db()
        self.assertEqual(p1.status, 'diterima')
        self.assertEqual(g.status, GelombangPPDB.DITUTUP)
        self.client.get(reverse('ppdb:ubah_status', args=[p2.pk, 'diterima']))
        p2.refresh_from_db()
        self.assertEqual(p2.status, 'cadangan')

    def test_filter_status_antrian(self):
        now = timezone.now()
        GelombangPPDB.objects.create(
            nama='Antrian',
            mulai=now - timedelta(days=1),
            selesai=now + timedelta(days=10),
            kuota=10,
            status=GelombangPPDB.DIBUKA,
        )
        self.client.post(reverse('pendaftaran'), _ppdb_payload(nama_lengkap='Berkas Uji'))
        p = Pendaftaran.objects.get(nama_lengkap='Berkas Uji')
        p.status = 'berkas_kurang'
        p.save()
        self.client.login(username='tu_kuota', password='sandi123')
        r = self.client.get(reverse('ppdb:antrian'), {'status': 'berkas_kurang'})
        self.assertContains(r, 'Berkas Uji')
        x = self.client.get(reverse('ppdb:antrian'), {'status': 'berkas_kurang', 'ekspor': 'xlsx'})
        self.assertEqual(x.status_code, 200)
        self.assertIn('spreadsheetml.sheet', x['Content-Type'])


class AkademikLanjutanTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.data = data_dasar()
        cls.tu = buat_user('tu_rekap', 'tata_usaha')
        cls.peg = Pegawai.objects.create(nama='Pengampu Rekap', jenis_kelamin='L')
        PengampuRB.objects.create(rb=cls.data['rb'], pegawai=cls.peg)
        cls.mapel = KitabAtauMapel.objects.create(unit=cls.data['unit'], nama='Nahwu', jenis='kitab')
        MapelRB.objects.create(rb=cls.data['rb'], mapel=cls.mapel)
        cls.santri = buat_santri('NISREK', 'Santri Rekap', '3601010101010200')
        KeanggotaanRB.objects.create(rb=cls.data['rb'], santri=cls.santri)

    def test_salin_rb(self):
        JadwalSlot.objects.create(
            rb=self.data['rb'], hari=0, jam_mulai=time(7, 0), jam_selesai=time(8, 0),
            mapel=self.mapel, pengampu=self.peg,
        )
        tahun_baru = TahunAjaran.objects.create(
            nama='2027/2028-uji',
            mulai=date.today() + timedelta(days=400),
            selesai=date.today() + timedelta(days=700),
        )
        baru = salin_rb(self.data['rb'], tahun_baru, salin_anggota=True)
        self.assertNotEqual(baru.pk, self.data['rb'].pk)
        self.assertEqual(baru.tahun_ajaran, tahun_baru)
        self.assertTrue(baru.pengampu.filter(pegawai=self.peg).exists())
        self.assertTrue(baru.mapel_rb.filter(mapel=self.mapel).exists())
        self.assertTrue(baru.anggota.filter(santri=self.santri).exists())
        self.assertTrue(baru.jadwal.exists())

    def test_rekap_alpa_beruntun(self):
        Pengaturan.get()
        p = Pengaturan.get()
        p.ambang_alpa = 3
        p.save()
        for i in range(3):
            pertemuan = Pertemuan.objects.create(
                rb=self.data['rb'], mapel=self.mapel, tanggal=date.today() - timedelta(days=i),
            )
            Absensi.objects.create(pertemuan=pertemuan, santri=self.santri, status='alpa')
        self.assertEqual(streak_alpa(self.santri), 3)
        self.client.login(username='tu_rekap', password='sandi123')
        r = self.client.get(reverse('akademik:rekap_absensi'))
        self.assertContains(r, 'Santri Rekap')
        self.assertContains(r, 'table-warning')


class IzinSantriTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        data_dasar()
        cls.user = buat_user('santri_izin', 'santri')
        cls.santri = buat_santri('NISIZN', 'Santri Izin Portal', '3601010101010300')
        cls.santri.user = cls.user
        cls.santri.save()

    def test_ajukan_izin_dari_portal(self):
        self.client.login(username='santri_izin', password='sandi123')
        now = timezone.now()
        r = self.client.post(reverse('kesiswaan:izin_santri'), {
            'jenis': 'pulang',
            'mulai': (now + timedelta(days=1)).strftime('%Y-%m-%dT%H:%M'),
            'selesai': (now + timedelta(days=2)).strftime('%Y-%m-%dT%H:%M'),
            'alasan': 'Keperluan keluarga',
        })
        self.assertEqual(r.status_code, 302)
        izin = Izin.objects.get(santri=self.santri)
        self.assertEqual(izin.status, 'diajukan')
        self.assertEqual(izin.jenis, 'pulang')
